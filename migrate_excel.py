"""
Excel 数据迁移脚本
将「每日工作-新1.xlsm」中的数据导入到 Work Log 系统数据库

运行方式：python migrate_excel.py <excel文件路径>
示例：python migrate_excel.py "每日工作-新1.xlsm"
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from datetime import date, datetime
from app.database import SessionLocal, engine, Base
from app.models import Department, User, Category, Project, DailyLog, TaskItem, BacklogItem
import openpyxl


# 类型映射：Excel类型 -> 系统work_type
WORK_TYPE_MAP = {
    'Bug': 'Bug', 'Report': 'Report', '功能': '功能',
    '增强': '增强', '接口': '接口', '学习': '学习',
    '培训': '培训', '打印': '打印', '问题排查': '问题排查',
}

# 状态映射：Excel状态 -> 系统status
STATUS_MAP = {
    '已完成': 'done',
    '未完成': 'doing',
    '挂起': 'blocked',
    '取消': 'cancelled',
}

# 任务分类映射
TASK_CATEGORY_MAP = {
    '运维工单': '运维工单', 'Bug工单': 'Bug工单', '临时任务': '临时任务',
    '迭代项': '迭代项', 'SRM 项目': 'SRM项目', '培训': '培训',
    '紧急插入临时任务': '紧急插入临时任务', '按需采购': '按需采购',
    '年度任务': '年度任务', '绩效': '绩效', 'STO项目': 'STO项目',
}


def ensure_tables():
    """确保所有表存在"""
    Base.metadata.create_all(bind=engine)
    print("✓ 数据库表已就绪")


def get_or_create_user(db: SessionLocal) -> User:
    """获取默认用户（zhengjie）"""
    user = db.query(User).filter(User.username == "zhengjie").first()
    if not user:
        dept = db.query(Department).first()
        if not dept:
            dept = Department(name="IT部")
            db.add(dept)
            db.flush()
        user = User(username="zhengjie", display_name="郑杰", dept_id=dept.id, role="admin")
        db.add(user)
        db.flush()
        dept.leader_id = user.id
        db.commit()
    return user


def get_or_create_category(db: SessionLocal, user_id: int, name: str) -> int:
    """获取或创建分类"""
    cat = db.query(Category).filter(Category.name == name, Category.user_id == user_id).first()
    if not cat:
        cat = Category(name=name, user_id=user_id, is_active=1)
        db.add(cat)
        db.flush()
    return cat.id


def migrate_daily_work(db: SessionLocal, user: User, ws):
    """迁移「每日工作」sheet"""
    print("\n📋 迁移「每日工作」...")
    count = 0
    skip = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        # 日期, 任务, 备注, 参考, 类型, 新增or修改, 事务码, 程序名, 接口, 增强, 类, 打印
        log_date, task_title, task_detail, reference, work_type, change_type, tcode, program_name, interface_name, enhancement_name, class_name, print_name = row

        if not log_date or not task_title:
            skip += 1
            continue

        # 转换日期
        if isinstance(log_date, datetime):
            log_date = log_date.date()
        elif isinstance(log_date, str):
            try:
                log_date = datetime.strptime(log_date, '%Y-%m-%d').date()
            except:
                skip += 1
                continue

        # 映射 work_type
        mapped_wt = WORK_TYPE_MAP.get(work_type, work_type) if work_type else None
        # 映射 change_type
        mapped_ct = change_type if change_type in ('新增', '修改') else None

        # 检查是否已存在（同一天同一标题）
        existing = db.query(DailyLog).filter(
            DailyLog.user_id == user.id,
            DailyLog.log_date == log_date,
            DailyLog.task_title == str(task_title),
        ).first()

        if existing:
            skip += 1
            continue

        # 根据类型自动匹配分类
        cat_name = None
        if work_type == 'Bug':
            cat_name = 'Bug工单'
        elif work_type in ('功能', '增强', 'Report', '接口'):
            cat_name = '开发'

        category_id = get_or_create_category(db, user.id, cat_name) if cat_name else None

        log = DailyLog(
            user_id=user.id,
            log_date=log_date,
            task_title=str(task_title),
            task_detail=str(task_detail) if task_detail else None,
            reference=str(reference) if reference else None,
            work_type=mapped_wt,
            change_type=mapped_ct,
            tcode=str(tcode).strip() if tcode else None,
            program_name=str(program_name).strip() if program_name else None,
            interface_name=str(interface_name).strip() if interface_name else None,
            enhancement_name=str(enhancement_name).strip() if enhancement_name else None,
            class_name=str(class_name).strip() if class_name else None,
            print_name=str(print_name).strip() if print_name else None,
            status='done',
            category_id=category_id,
        )
        db.add(log)
        count += 1

    db.commit()
    print(f"  ✓ 导入 {count} 条，跳过 {skip} 条")


def migrate_tasks(db: SessionLocal, user: User, ws):
    """迁移「任务项」sheet"""
    print("\n✅ 迁移「任务项」...")
    count = 0
    skip = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        # 任务, 截止日期, 任务分类, 完成状态, 备注, 剩余日期
        task_title, deadline, task_category, status, notes, _ = row

        if not task_title:
            skip += 1
            continue

        # 转换截止日期
        if isinstance(deadline, datetime):
            deadline = deadline.date()
        elif isinstance(deadline, str) and deadline:
            try:
                deadline = datetime.strptime(deadline, '%Y-%m-%d').date()
            except:
                deadline = None
        else:
            deadline = None

        # 映射分类
        mapped_cat = TASK_CATEGORY_MAP.get(task_category, task_category) if task_category else None
        # 映射状态
        mapped_status = status if status in ('已完成', '未完成', '挂起', '取消') else '未完成'

        # 检查是否已存在
        existing = db.query(TaskItem).filter(
            TaskItem.user_id == user.id,
            TaskItem.task_title == str(task_title),
        ).first()

        if existing:
            skip += 1
            continue

        task = TaskItem(
            user_id=user.id,
            task_title=str(task_title),
            deadline=deadline,
            task_category=mapped_cat,
            status=mapped_status,
            notes=str(notes) if notes else None,
        )
        db.add(task)
        count += 1

    db.commit()
    print(f"  ✓ 导入 {count} 条，跳过 {skip} 条")


def migrate_backlog(db: SessionLocal, user: User, ws):
    """迁移「待开发项」sheet"""
    print("\n📋 迁移「待开发项」...")
    count = 0
    skip = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        # 任务, 截止日期, 剩余天数, 待完成任务数
        task_title, deadline, _, pending_count = row

        if not task_title:
            skip += 1
            continue

        # 转换截止日期
        if isinstance(deadline, datetime):
            deadline = deadline.date()
        elif isinstance(deadline, str) and deadline:
            try:
                deadline = datetime.strptime(deadline, '%Y-%m-%d').date()
            except:
                deadline = None
        else:
            deadline = None

        # 检查是否已存在
        existing = db.query(BacklogItem).filter(
            BacklogItem.user_id == user.id,
            BacklogItem.task_title == str(task_title),
        ).first()

        if existing:
            skip += 1
            continue

        item = BacklogItem(
            user_id=user.id,
            task_title=str(task_title),
            deadline=deadline,
            pending_count=int(pending_count) if pending_count else None,
        )
        db.add(item)
        count += 1

    db.commit()
    print(f"  ✓ 导入 {count} 条，跳过 {skip} 条")


def main():
    if len(sys.argv) < 2:
        print("用法：python migrate_excel.py <excel文件路径>")
        print("示例：python migrate_excel.py \"每日工作-新1.xlsm\"")
        sys.exit(1)

    excel_path = sys.argv[1]
    if not os.path.exists(excel_path):
        print(f"文件不存在：{excel_path}")
        sys.exit(1)

    print(f"📂 读取 Excel：{excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    ensure_tables()

    db = SessionLocal()
    try:
        user = get_or_create_user(db)
        print(f"👤 用户：{user.display_name} (ID: {user.id})")

        # 迁移各 sheet
        if '每日工作' in wb.sheetnames:
            migrate_daily_work(db, user, wb['每日工作'])
        else:
            print("⚠️ 未找到「每日工作」sheet")

        if '任务项' in wb.sheetnames:
            migrate_tasks(db, user, wb['任务项'])
        else:
            print("⚠️ 未找到「任务项」sheet")

        if '待开发项' in wb.sheetnames:
            migrate_backlog(db, user, wb['待开发项'])
        else:
            print("⚠️ 未找到「待开发项」sheet")

        print("\n🎉 迁移完成！")

    except Exception as e:
        db.rollback()
        print(f"\n❌ 迁移失败：{e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()


if __name__ == "__main__":
    main()
